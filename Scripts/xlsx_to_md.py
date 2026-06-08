#!/usr/bin/env python3
"""Convert Excel (.xlsx) to Markdown."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from Tools.drawio_writer import from_json, write_drawio
from Tools.markdown_utils import build_frontmatter, ensure_output_dir, table_to_gfm


def _sheet_to_rows(sheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    # Trim trailing empty rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def convert(input_path: Path, output_dir: Path, export_charts: bool = True) -> Path:
    from openpyxl import load_workbook

    input_path = Path(input_path)
    basename = input_path.stem
    out_dir = ensure_output_dir(output_dir, basename)
    charts_dir = out_dir / "charts"
    if export_charts:
        charts_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        parts.append(f"## {sheet_name}")
        parts.append("")
        rows = _sheet_to_rows(sheet)
        if rows:
            parts.append(table_to_gfm(rows))
        else:
            parts.append("*Empty sheet*")
        parts.append("")

        # Export chart placeholders when charts exist on sheet
        if export_charts and hasattr(sheet, "_charts") and sheet._charts:
            for idx, chart in enumerate(sheet._charts, start=1):
                chart_name = f"{sheet_name}_chart{idx}"
                chart_json = {
                    "name": chart_name,
                    "nodes": [
                        {"key": "chart", "label": f"Chart: {chart_name}\n(Edit in draw.io)"},
                    ],
                    "edges": [],
                }
                drawio_path = charts_dir / f"{chart_name}.drawio"
                write_drawio(from_json(chart_json), drawio_path)
                parts.append(f"[Edit chart in draw.io](charts/{chart_name}.drawio)")
                parts.append("")

    md_path = out_dir / f"{basename}.md"
    frontmatter = build_frontmatter(
        title=basename,
        source=str(input_path),
        converter="xlsx_to_md",
    )
    md_path.write_text(frontmatter + "\n".join(parts), encoding="utf-8")
    return md_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel to Markdown")
    parser.add_argument("--input", "-i", required=True)
    parser.add_argument("--output", "-o", default="output")
    parser.add_argument("--no-charts", action="store_true")
    args = parser.parse_args()

    try:
        out = convert(Path(args.input), Path(args.output), export_charts=not args.no_charts)
        print(out)
        return 0
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
